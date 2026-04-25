"""Excel-backed Q&A tools for shared DataSheriff knowledge."""

from __future__ import annotations

from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
import re
from threading import Lock
from typing import Any

from mcp.server.fastmcp import FastMCP

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - handled at runtime via tool output
    load_workbook = None  # type: ignore[assignment]

try:
    from mcp_server.config import get_settings
except ModuleNotFoundError:
    from config import get_settings


@dataclass
class _Cache:
    path: Path | None = None
    mtime: float = -1.0
    rows: list[dict[str, Any]] | None = None


_CACHE = _Cache()
_CACHE_LOCK = Lock()


def _normalize(text: str) -> str:
    cleaned = re.sub(r"[^a-z0-9\s]", " ", (text or "").lower())
    return " ".join(cleaned.split())


def _similarity(query: str, candidate: str) -> float:
    q = _normalize(query)
    c = _normalize(candidate)
    if not q or not c:
        return 0.0

    ratio = SequenceMatcher(None, q, c).ratio()
    q_tokens = set(q.split())
    c_tokens = set(c.split())
    overlap = len(q_tokens & c_tokens) / max(1, len(q_tokens | c_tokens))
    return (0.7 * ratio) + (0.3 * overlap)


def _read_excel_rows() -> tuple[list[dict[str, Any]], str | None]:
    settings = get_settings()
    path_raw = (settings.excel_kb_path or "").strip()
    if not path_raw:
        return [], "EXCEL_KB_PATH is not configured."
    if load_workbook is None:
        return [], "Missing dependency 'openpyxl'. Install requirements and restart."

    path = Path(path_raw).expanduser()
    if not path.exists():
        return [], f"Excel knowledge file not found at: {path}"

    with _CACHE_LOCK:
        mtime = path.stat().st_mtime
        if _CACHE.path == path and _CACHE.rows is not None and _CACHE.mtime == mtime:
            return _CACHE.rows, None

        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            if settings.excel_kb_sheet in workbook.sheetnames:
                sheet = workbook[settings.excel_kb_sheet]
            else:
                sheet = workbook.active

            values = list(sheet.iter_rows(values_only=True))
            if not values:
                return [], "Excel sheet is empty."

            header = [str(cell or "").strip().lower() for cell in values[0]]
            q_col = settings.excel_kb_question_column.strip().lower()
            a_col = settings.excel_kb_answer_column.strip().lower()
            k_col = settings.excel_kb_keywords_column.strip().lower()

            if q_col not in header or a_col not in header:
                return [], (
                    "Missing required columns in Excel. "
                    f"Expected '{q_col}' and '{a_col}'. Found: {', '.join(header)}"
                )

            q_idx = header.index(q_col)
            a_idx = header.index(a_col)
            k_idx = header.index(k_col) if k_col in header else None

            parsed_rows: list[dict[str, Any]] = []
            for row_idx, row in enumerate(values[1:], start=2):
                question = str(row[q_idx] or "").strip() if q_idx < len(row) else ""
                answer = str(row[a_idx] or "").strip() if a_idx < len(row) else ""
                if not question or not answer:
                    continue

                keywords_raw = ""
                if k_idx is not None and k_idx < len(row):
                    keywords_raw = str(row[k_idx] or "")
                keywords = [part.strip() for part in keywords_raw.split(",") if part.strip()]

                parsed_rows.append(
                    {
                        "row": row_idx,
                        "question": question,
                        "answer": answer,
                        "keywords": keywords,
                    }
                )

            _CACHE.path = path
            _CACHE.mtime = mtime
            _CACHE.rows = parsed_rows
            return parsed_rows, None
        finally:
            workbook.close()


def register_excel_kb_tools(mcp: FastMCP) -> None:
    """Register optional Excel-backed tools."""

    @mcp.tool()
    def answer_from_excel(question: str, min_score: float = 0.45) -> str:
        """
        Answer a user question from a shared Excel knowledge base.
        The Excel file is configured via EXCEL_KB_PATH and should contain:
        - question (required)
        - answer (required)
        - keywords (optional, comma-separated)

        Returns an answer when a confident match is found.
        Returns NO_MATCH when no row clears min_score.
        """
        rows, error = _read_excel_rows()
        if error:
            return f"Excel KB unavailable: {error}"
        if not rows:
            return "Excel KB has no usable rows."

        best: dict[str, Any] | None = None
        best_score = 0.0
        for row in rows:
            q_score = _similarity(question, str(row.get("question") or ""))
            keyword_text = " ".join(row.get("keywords") or [])
            k_score = _similarity(question, keyword_text) if keyword_text else 0.0
            score = max(q_score, k_score)
            if score > best_score:
                best_score = score
                best = row

        if not best or best_score < min_score:
            return f"NO_MATCH: no spreadsheet row matched confidently (best_score={best_score:.2f})."

        return (
            f"{best['answer']}\n\n"
            f"Source: Excel row {best['row']} (match_score={best_score:.2f})"
        )

    @mcp.tool()
    def list_excel_questions(limit: int = 25) -> str:
        """
        List available questions from the configured Excel knowledge base.
        Useful to discover what users can ask the Slack bot.
        """
        rows, error = _read_excel_rows()
        if error:
            return f"Excel KB unavailable: {error}"
        if not rows:
            return "Excel KB has no usable rows."

        capped = rows[: max(1, min(limit, 200))]
        lines = [f"Excel KB contains {len(rows)} row(s). Showing {len(capped)}:"]
        for idx, row in enumerate(capped, start=1):
            lines.append(f"{idx}. {row['question']}")
        return "\n".join(lines)
